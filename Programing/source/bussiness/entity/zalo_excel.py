import Programing.source.utility.fix_qt_import_error
import openpyxl
from PyQt5 import QtWidgets, QtCore

from Programing.source.bussiness.controller.zalo_controller import *


class Excel:

    def __init__(self):
        self.nicks_path = ''
        self.status_summary = Path.STATUS_PATH

    def read_excel(self):
        """ read assets and validate assets from excel file """
        book = openpyxl.load_workbook(self.nicks_path)
        sheet = book[book.sheetnames[0]]
        data = [[c1.value, c2.value, c3.value] for c1, c2, c3 in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=3)
                if c1.value is not None]
        return data

    def get_excel_file(self, title="Excel files", directory='.'):
        """ locate folder containning excel file """
        widget = QtWidgets.QWidget(flags=QtCore.Qt.WindowStaysOnTopHint)
        filename, filter_ = QtWidgets.QFileDialog.getOpenFileName(
            parent=widget,
            directory=directory,
            caption=title,
            filter='Excel files (*.xlsx)'
        )
        return filename

    def write_status_to_excel(self, data):
        """ write status after sending and getting status"""
        book = openpyxl.Workbook()
        sheet = book.active
        for row in data:
            sheet.append(row)
        Excel.format_sheet(sheet)
        book.save(Path.STATUS_PATH)

    @staticmethod
    def format_sheet(sheet):
        """ format excel file """
        # bold first row
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = cell.font.copy(bold=True)

        # resize column width with length of largest cell
        for column_cells in sheet.columns:
            length = max(len(Excel.as_text(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column].width = length

    @staticmethod
    def as_text(value):
        """ convert arbitrary value to string """
        return str(value) if value is not None else ""

    def open_status(self):
        """ open file contains message's status """
        try:
            os.startfile(self.status_summary)
        except FileNotFoundError:
            ZaloUtility.show_message("THÔNG BÁO", "<center> Hiện chưa có file tổng hợp tin nhắn </center>")
        except Exception:
            pass

    @staticmethod
    def take_phones_for_updating():
        """ take the phone number of message which have status differ 'ĐÃ XEM' """
        book = openpyxl.load_workbook(Path.STATUS_PATH)
        sheet = book[book.sheetnames[0]]
        phone_numbers = []
        filter_conditions = ["CHẶN TÔI", "CHẶN TIN NHẮN TỪ NGƯỜI LẠ", "KHÔNG TÌM THẤY NICK ZALO"]
        for row in sheet.iter_rows(min_col=1, max_col=1, min_row=2, max_row=sheet.max_row):
            if str(sheet['G' + str(row[0].row)].value).upper() not in filter_conditions and\
               Excel.as_text(sheet['E' + str(row[0].row)].value).upper() not in ["ĐÃ XEM", ""]:

                phone_numbers.append(row[0].value)
        return phone_numbers

    def gui_browse_nicks(self):
        path = self.get_excel_file('Chọn file chứa nick Zalo')
        if path != "":
            self.nicks_path = path

