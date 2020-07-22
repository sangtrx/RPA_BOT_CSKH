import os

import Programing.source.utility.fix_qt_import_error
import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
from selenium.common.exceptions import NoSuchElementException, NoSuchWindowException, WebDriverException

from Programing.source.bussiness.controller.zalo_utils import ZaloUtility, Path
from Programing.source.bussiness.entity.zalo_excel import Excel
from Programing.source.bussiness.controller.zalo_controller import Zalo


def remove_qt_temporary_files():
    if os.path.exists('qt.conf'):
        os.remove('qt.conf')


class Ui_Form:

    def __init__(self, Form):
        self.excel = Excel()
        self.zalo = Zalo()

        self.make_folder()

        self.excel.nicks_path = ZaloUtility.get_saved_paths(file=self.zalo.saved_path)
        self.qss_path = ZaloUtility.get_resource_path("stylesheet.qss")

        self.setupUi(Form)
        if self.lineEdit_nick_zalo.text() == '':
            self.lineEdit_nick_zalo.setPlaceholderText('Chọn và import file')

    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(850, 544)
        Form.setStyleSheet(open(self.qss_path, mode='r', encoding='utf-8').read())
        self.frame_main = QtWidgets.QFrame(Form, flags=QtCore.Qt.WindowStaysOnTopHint)
        self.frame_main.setGeometry(QtCore.QRect(10, 0, 831, 541))
        self.frame_main.setStyleSheet("")
        self.frame_main.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_main.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_main.setObjectName("frame_main")

        self.frame_label_image = QtWidgets.QFrame(self.frame_main, flags=QtCore.Qt.WindowStaysOnTopHint)
        self.frame_label_image.setGeometry(QtCore.QRect(20, 10, 781, 231))
        self.frame_label_image.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_label_image.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_label_image.setObjectName("frame_label_image")

        self.label_cskh = QtWidgets.QLabel(self.frame_label_image)
        self.label_cskh.setGeometry(QtCore.QRect(30, 180, 741, 41))
        font = QtGui.QFont()
        font.setFamily("Roboto")
        font.setItalic(False)
        self.label_cskh.setFont(font)
        self.label_cskh.setStyleSheet("selection-color: rgb(255, 0, 0);")
        self.label_cskh.setAlignment(QtCore.Qt.AlignCenter)
        self.label_cskh.setObjectName("label_cskh")

        self.image_AC_solution = QtWidgets.QLabel(self.frame_label_image)
        self.image_AC_solution.setGeometry(QtCore.QRect(280, 10, 211, 161))
        self.image_AC_solution.setText("")
        self.image_AC_solution.setPixmap(QtGui.QPixmap(Path.IMG_PATH))
        self.image_AC_solution.setScaledContents(True)
        self.image_AC_solution.setAlignment(QtCore.Qt.AlignCenter)
        self.image_AC_solution.setObjectName("image_AC_solution")

        self.frame_btn_gui_tn = QtWidgets.QFrame(self.frame_main, flags=QtCore.Qt.WindowStaysOnTopHint)
        self.frame_btn_gui_tn.setGeometry(QtCore.QRect(30, 430, 771, 80))
        self.frame_btn_gui_tn.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_btn_gui_tn.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_btn_gui_tn.setObjectName("frame_btn_gui_tn")

        self.button_gui_tn = QtWidgets.QPushButton(self.frame_btn_gui_tn)
        self.button_gui_tn.setGeometry(QtCore.QRect(630, 20, 131, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.button_gui_tn.setFont(font)
        self.button_gui_tn.setStyleSheet("")
        self.button_gui_tn.setObjectName("button_gui_tn")
        self.button_gui_tn.clicked.connect(self.send_message)

        self.button_xem_file = QtWidgets.QPushButton(self.frame_btn_gui_tn)
        self.button_xem_file.setGeometry(QtCore.QRect(150, 20, 161, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.button_xem_file.setFont(font)
        self.button_xem_file.setStyleSheet("")
        self.button_xem_file.setObjectName("button_xem_file")
        self.button_xem_file.clicked.connect(self.excel.open_status)

        self.button_cap_nhat_trang_thai = QtWidgets.QPushButton(self.frame_btn_gui_tn)
        self.button_cap_nhat_trang_thai.setGeometry(QtCore.QRect(320, 20, 171, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.button_cap_nhat_trang_thai.setFont(font)
        self.button_cap_nhat_trang_thai.setStyleSheet("")
        self.button_cap_nhat_trang_thai.setObjectName("button_cap_nhat_trang_thai")
        self.button_cap_nhat_trang_thai.clicked.connect(self.update_status)

        self.frame_btn_choose = QtWidgets.QFrame(self.frame_main)
        self.frame_btn_choose.setGeometry(QtCore.QRect(650, 310, 151, 111))
        self.frame_btn_choose.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_btn_choose.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_btn_choose.setObjectName("frame_btn_choose")

        self.button_nick_zalo = QtWidgets.QPushButton(self.frame_btn_choose)
        self.button_nick_zalo.setGeometry(QtCore.QRect(10, 50, 131, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.button_nick_zalo.setFont(font)
        self.button_nick_zalo.setStyleSheet("")
        self.button_nick_zalo.setObjectName("button_nick_zalo")
        self.button_nick_zalo.clicked.connect(self.choose_zalo_nick)

        self.frame_lineEdit_file = QtWidgets.QFrame(self.frame_main, flags=QtCore.Qt.WindowStaysOnTopHint)
        self.frame_lineEdit_file.setGeometry(QtCore.QRect(30, 310, 611, 111))
        self.frame_lineEdit_file.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_lineEdit_file.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_lineEdit_file.setObjectName("frame_lineEdit_file")

        self.lineEdit_nick_zalo = QtWidgets.QLineEdit(self.frame_lineEdit_file)
        self.lineEdit_nick_zalo.setGeometry(QtCore.QRect(20, 50, 581, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.lineEdit_nick_zalo.setFont(font)
        self.lineEdit_nick_zalo.setStyleSheet("")
        self.lineEdit_nick_zalo.setText("")
        self.lineEdit_nick_zalo.setObjectName("lineEdit_nick_zalo")
        self.lineEdit_nick_zalo.setReadOnly(True)
        self.lineEdit_nick_zalo.setText(ZaloUtility.get_saved_paths(self.zalo.saved_path))

        self.label_nick_zalo = QtWidgets.QLabel(self.frame_lineEdit_file)
        self.label_nick_zalo.setGeometry(QtCore.QRect(20, 20, 321, 21))
        font = QtGui.QFont()
        font.setFamily("Roboto")
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_nick_zalo.setFont(font)
        self.label_nick_zalo.setObjectName("label_nick_zalo")

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "CHĂM SÓC KHÁCH HÀNG TỰ ĐỘNG"))
        self.label_cskh.setText(_translate("Form", "<html><head/><body><p>GIẢI PHÁP CHĂM SÓC KHÁCH HÀNG TỰ ĐỘNG</p></body></html>"))
        self.button_gui_tn.setText(_translate("Form", "Gửi tin nhắn"))
        self.button_xem_file.setText(_translate("Form", "Xem trạng thái"))
        self.button_cap_nhat_trang_thai.setText(_translate("Form", "Cập nhật trạng thái"))
        self.button_nick_zalo.setToolTip(_translate("Form", "<html><head/><body><p><br/></p></body></html>"))
        self.button_nick_zalo.setText(_translate("Form", "Chọn file"))
        self.label_nick_zalo.setText(_translate("Form", "Chọn file chứa nick Zalo"))

    def choose_zalo_nick(self):
        self.excel.gui_browse_nicks()
        ZaloUtility.save_paths(source_file=self.zalo.saved_path, new_path=self.excel.nicks_path)
        self.lineEdit_nick_zalo.setText(self.excel.nicks_path) if self.excel.nicks_path else None

    def send_message(self):
        if self.excel.nicks_path == "":
            ZaloUtility.show_message('CẢNH BÁO', '<center> Bạn chưa chọn file excel chứa tin nhắn và ảnh </center>', type_=QMessageBox.Critical)
            return

        try:
            excel_data = self.excel.read_excel()
        except FileNotFoundError:
            ZaloUtility.show_message("LỖI", f" File <font color='green' size=4><b><i> {self.excel.nicks_path} </i></b></font> không tồn tại <br>"
                                     "Bạn vui lòng kiểm tra lại đường dẫn file", type_=QMessageBox.Critical)
            return

        try:
            self.zalo.login()
        except NoSuchElementException:
            return
        if not self.zalo.login_ok:
            return

        for data in excel_data:
            nick, message, image_folder = data[0], data[1], data[2]
            if nick is not None:
                self.zalo.find_contact(nick)
                self.zalo.phones.append(nick)
                if self.zalo.contact_found:
                    self.zalo.send_data(message, image_folder)
                    self.zalo.found_phones.append(nick)
            else:
                pass
        self.take_and_log_status()
        ZaloUtility.show_message("THÔNG BÁO", "<center> Đã gửi tin nhắn xong </center> \n"
                                 "<center> Bấm nút <font size=4 color='purple'><b> Xem trạng thái </b></font> "
                                 "để xem trạng thái các tin đã gửi </center> ", type_=QMessageBox.Information)

    def take_and_log_status(self):
        # get and log status of the found phones
        excel_datas = self.excel.read_excel()
        for row in excel_datas:
            status_datas = self.zalo.take_message_status(row[0])
            row.extend(status_datas)
        df = [Zalo.STATUS_FIELDS]
        df.extend(excel_datas)
        self.excel.write_status_to_excel(df)

    def update_status(self):
        """update everyone's status whose statuses are 'ĐÃ GỬI'"""
        # check whether the users choose file or not
        try:
            book = openpyxl.load_workbook(Path.STATUS_PATH)
        except FileNotFoundError:
            ZaloUtility.show_message("THÔNG BÁO", "Không có file trạng thái để cập nhật", type_=QMessageBox.Information)
            return

        # check whether the users close 'status_summary.xlsx' file or not
        try:
            book.save(Path.STATUS_PATH)
        except PermissionError:
            ZaloUtility.show_message("LỖI", "Bạn vui lòng đóng file <font color='green' size=4><b><i> status_summary.xlsx </i></b></font> "
                                     "<br> và thực hiện lại",  type_=QMessageBox.Critical)
            return
        sheet = book[book.sheetnames[0]]
        phone_numbers = Excel.take_phones_for_updating()
        if not phone_numbers:
            ZaloUtility.show_message('THÔNG BÁO', '<center> Các tin nhắn đều có trạng thái'
                                     ' <font color="blue"><b><i>ĐÃ XEM </font></b></i> </center>\n'
                                     '<center> Nhấn vào nút <font size=4 color="purple"><b> Xem trạng thái </b></font> để kiểm tra lại </center>',
                                     type_=QMessageBox.Information)
            return
        try:
            self.zalo.login()
        except (WebDriverException, NoSuchWindowException):
            return
        phone_and_status = dict()
        for phone in phone_numbers:
            try:
                status = self.zalo.update_status_nick(phone)
            except (WebDriverException, NoSuchWindowException, NoSuchElementException):
                return
            phone_and_status[phone] = status

        # write status after updating to corresponding phone
        for row in sheet.iter_rows(min_col=1, max_col=1, min_row=2, max_row=sheet.max_row):
            if Excel.as_text(row[0].value) in phone_and_status.keys():
                sheet['E' + Excel.as_text(row[0].row)].value = phone_and_status[Excel.as_text(row[0].value)]
        book.save(Path.STATUS_PATH)
        ZaloUtility.show_message("THÔNG BÁO", "<center> Đã cập nhật xong trạng thái </center>\n"
                                 "<center> Bấm nút <font size=4 color='purple'><b> Xem trạng thái </b></font> để xem lại cập nhật </center>"
                                 , type_=QMessageBox.Information)

    def make_folder(self):
        """ create a AC_SOLUTION's folder that saved 'saved_paths.txt' file """
        if not os.path.exists(Path.BOT_ROOT_FOLDER):
            os.makedirs(Path.BOT_ROOT_FOLDER)
            if not os.path.isfile(self.zalo.saved_path):
                open(self.zalo.saved_path, mode="w+")


